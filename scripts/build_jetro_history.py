# -*- coding: utf-8 -*-
"""
JETRO「投資関連コスト比較調査」の過去年度分からワーカー賃金の時系列を作り
data/jetro_history.json に書き出す（年1回のワークフローで全再構築）。

- 一覧ページの全レポートリンクをたどり、タイトルから「20XX年度」を読み取る
- Excel添付がある年度（おおむね2017年度以降）だけを対象にする
- 都市キーは「都市名|国名」。同一年度に複数調査で同じ都市が出た場合は先勝ち
"""
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone

import requests

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
UA = {"User-Agent": "Mozilla/5.0 (compatible; procurement-cost-site)"}
LIST_URL = "https://www.jetro.go.jp/world/business_environment/cost.html"


def worker_usd_from_sheet(ws):
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=6):
        cells = [c.value for c in row]
        texts = [str(c) if c is not None else "" for c in cells]
        if "ワーカー（一般工職）" in texts[3]:
            v = cells[4]
            if isinstance(v, (int, float)):
                return round(float(v), 1)
            try:
                return round(float(str(v).replace(",", "").strip()), 1)
            except (ValueError, AttributeError):
                return None
    return None


def main():
    import openpyxl
    r = requests.get(LIST_URL, timeout=120, headers=UA)
    r.encoding = r.apparent_encoding
    hrefs = []
    for h in re.findall(r'href="(/world/reports/[^"]+\.html)[^"]*"', r.text):
        if h not in hrefs:
            hrefs.append(h)
    print("レポートリンク:", len(hrefs))

    cities = {}     # "都市|国" -> {"city","country","series":{年度: usd}}
    surveys = []
    for h in hrefs:
        url = "https://www.jetro.go.jp" + h
        try:
            rp = requests.get(url, timeout=120, headers=UA)
            rp.encoding = rp.apparent_encoding
            tm = re.search(r"<title>([^<|]+)", rp.text)
            title = tm.group(1).strip() if tm else ""
            ym = re.search(r"(20\d{2})年度", title)
            if not ym or "コスト比較" not in title:
                continue
            year = int(ym.group(1))
            xm = re.search(r'href="([^"]+\.xlsx?)"', rp.text)
            if not xm:
                continue
            xurl = xm.group(1)
            if xurl.startswith("/"):
                xurl = "https://www.jetro.go.jp" + xurl
            xr = requests.get(xurl, timeout=300, headers=UA)
            xr.raise_for_status()
            wb = openpyxl.load_workbook(io.BytesIO(xr.content), data_only=True)
            n = 0
            for name in wb.sheetnames:
                m = re.match(r"^\s*(.+?)\s*[（(]\s*(.+?)\s*[）)]\s*$", name)
                if not m or name.strip() in ("概要", "目次"):
                    continue
                city, country = m.group(1).strip(), m.group(2).strip()
                usd = worker_usd_from_sheet(wb[name])
                if usd is None:
                    continue
                key = f"{city}|{country}"
                rec = cities.setdefault(key, {"city": city, "country": country, "series": {}})
                if year not in rec["series"]:
                    rec["series"][year] = usd
                    n += 1
            surveys.append({"year": year, "label": title, "page": url})
            print(f"  {year}年度 | {title[:44]} -> {n}都市")
        except Exception as e:
            print(f"  SKIP {url}: {str(e)[:60]}")

    years_all = sorted({y for c in cities.values() for y in c["series"]})
    # 3時点以上ある都市だけ残す（時系列として意味を持たせる）
    kept = {k: v for k, v in cities.items() if len(v["series"]) >= 3}
    out = {
        "updated": datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d"),
        "source": "ジェトロ「投資関連コスト比較調査」各年度（ワーカー（一般工職）月額賃金・米ドル）",
        "years": years_all,
        "cities": kept,
    }
    with open(os.path.join(DATA, "jetro_history.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    print(f"jetro_history.json: {len(kept)}都市 × 年度{years_all}")
    if len(kept) < 30:
        sys.exit("ERROR: 時系列都市が少なすぎます")


if __name__ == "__main__":
    main()
