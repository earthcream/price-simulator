# -*- coding: utf-8 -*-
"""
JETRO「投資関連コスト比較調査」から海外都市のワーカー賃金等を取得し data/jetro.json に書き出す。

仕組み（年1回更新・自動発見）:
  1. 一覧ページ https://www.jetro.go.jp/world/business_environment/cost.html から
     各地域（東アジア、アジア大洋州・日本、欧州、北米、中南米、中東、アフリカ）の
     最新レポートページURLを拾う（一覧の最初に出てくるリンク＝最新年度）
  2. 各レポートページから .xlsx を見つけてダウンロード
  3. 都市別シートから以下を抽出
     - ワーカー（一般工職）月額賃金（米ドル）
     - エンジニア（中堅技術者）月額賃金（米ドル）
     - 中間管理職（課長クラス）月額賃金（米ドル）
     - 名目賃金上昇率（直近3年のテキスト）
     - 調査実施時期・換算レート

JETROがページ構成を変えると失敗する。失敗時はGitHub Actionsの通知が届くので
docs/年次データの更新手順.md の手順で手動対応する。
"""
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone

import requests

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
UA = {"User-Agent": "Mozilla/5.0 (compatible; procurement-cost-site)"}
LIST_URL = "https://www.jetro.go.jp/world/business_environment/cost.html"
REGIONS = ["東アジア", "アジア大洋州・日本", "欧州", "北米", "中南米", "中東", "アフリカ"]

WAGE_ROWS = {
    "worker": "ワーカー（一般工職）",
    "engineer": "エンジニア（中堅技術者）",
    "manager": "中間管理職（課長クラス）",
}


def discover_reports():
    """一覧ページから 地域名 -> 最新レポートURL を取る"""
    r = requests.get(LIST_URL, timeout=120, headers=UA)
    r.encoding = r.apparent_encoding
    # (href, リンクテキスト) を順に取り、地域名ごとに最初の /world/reports/ リンクを採用
    pairs = re.findall(r'<a[^>]+href="(/world/reports/[^"]+)"[^>]*>([^<]{1,40})</a>', r.text)
    found = {}
    for href, text in pairs:
        t = text.strip()
        if t in REGIONS and t not in found:
            found[t] = "https://www.jetro.go.jp" + href.split("?")[0]
    missing = [x for x in REGIONS if x not in found]
    if missing:
        sys.exit(f"ERROR: 一覧ページから地域が見つかりません: {missing}（ページ構成変更の可能性）")
    return found


def find_xlsx(report_url):
    r = requests.get(report_url, timeout=120, headers=UA)
    r.encoding = r.apparent_encoding
    title = re.search(r"<title>([^<|]+)", r.text)
    m = re.search(r'href="([^"]+\.xlsx?)"', r.text)
    if not m:
        sys.exit(f"ERROR: {report_url} にExcelが見つかりません")
    url = m.group(1)
    if url.startswith("/"):
        url = "https://www.jetro.go.jp" + url
    return url, (title.group(1).strip() if title else report_url)


def num_or_none(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 1)
    s = str(v).replace(",", "").strip()
    try:
        return round(float(s), 1)
    except ValueError:
        return None


def parse_xlsx(content, region):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    cities = []
    for name in wb.sheetnames:
        m = re.match(r"^\s*(.+?)\s*[（(]\s*(.+?)\s*[）)]\s*$", name)
        if not m or name.strip() in ("概要", "目次"):
            continue
        city, country = m.group(1).strip(), m.group(2).strip()
        ws = wb[name]
        rec = {"city": city, "country": country, "region": region,
               "survey": "", "rate": "", "growth": ""}
        for row in ws.iter_rows(min_row=1, max_row=60, max_col=6):
            cells = [c.value for c in row]
            texts = [str(c) if c is not None else "" for c in cells]
            joined = " ".join(texts)
            if "調査実施時期" in joined and not rec["survey"]:
                rec["survey"] = re.sub(r".*調査実施時期[：:]\s*", "", joined).strip()[:40]
            if "換算レート" in joined and not rec["rate"]:
                rec["rate"] = re.sub(r".*換算レート[：:]\s*", "", joined).strip()[:60]
            label = texts[3]
            for key, kw in WAGE_ROWS.items():
                if kw in label and key not in rec:
                    pass
            for key, kw in WAGE_ROWS.items():
                if kw in label:
                    rec[key + "_usd"] = num_or_none(cells[4])
                    if rec[key + "_usd"] is None:
                        rec[key + "_raw"] = texts[4].strip()[:60]
            if "名目賃金上昇率" in label:
                rec["growth"] = texts[4].strip()[:120]
        if "worker_usd" in rec or "worker_raw" in rec:
            cities.append(rec)
    return cities


def main():
    reports = discover_reports()
    all_cities, sources = [], {}
    for region, url in reports.items():
        xlsx_url, title = find_xlsx(url)
        r = requests.get(xlsx_url, timeout=300, headers=UA)
        r.raise_for_status()
        cities = parse_xlsx(r.content, region)
        print(f"{region}: {title[:40]} -> {len(cities)}都市")
        if len(cities) < 3:
            sys.exit(f"ERROR: {region} の都市数が少なすぎます（様式変更の可能性）")
        all_cities += cities
        sources[region] = {"label": title, "page": url, "xlsx": xlsx_url}
    out = {
        "updated": datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d"),
        "source": "ジェトロ（日本貿易振興機構）「投資関連コスト比較調査」",
        "regions": sources,
        "cities": all_cities,
    }
    with open(os.path.join(DATA, "jetro.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    n_worker = sum(1 for c in all_cities if c.get("worker_usd") is not None)
    print(f"jetro.json: {len(all_cities)}都市（うちワーカー賃金の数値あり {n_worker}）")


if __name__ == "__main__":
    main()
